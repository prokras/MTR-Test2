import requests
import sys
import props
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('metadata.xlsx', data_only=True)
print(wb.get_sheet_names())
ws = wb.get_sheet_by_name('test_meta')
print(ws.title)
# print(tuple(ws.rows))
for row in ws.iter_rows():
    print(row[0].value)

driver = webdriver.Chrome(props.chrome_path)
driver.get("http://magentofinal.mytriorings.com")
print(driver.title)
driver.quit()
